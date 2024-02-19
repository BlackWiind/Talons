import re

from openpyxl import load_workbook

from .models import Services, IdDepartments


def update_db_from_excel(file):
    """
    Создаёт записи в бд на основе входящего excel файла.
    :param file:
    :return:
    """
    wb = load_workbook(file)
    ws = wb.active
    Services.objects.all().delete()
    id_values = IdDepartments.objects.all()
    name: str
    price: str
    service_type: str
    # row - список ячеек из строки файла.
    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            # берём id в первой ячейке
            row_id = int(reg_exp(row[1]))
            # проверяем содержит ли строка услуги
            if row[7] != '':
                name = row[5]  # название услуги из второй ячейки
                service_type = row[6]  # тип услуги из третьей
                price = round(row[7], 2)  # округлённый прайс из четвёртой ячейки
                for dep in id_values.filter(id_value=row_id):
                    Services.objects.create(name=name,
                                            price=price,
                                            type=service_type,
                                            department=dep.department)
        except Exception as e:
            print(f'{reg_exp(row[1])}   {e}')


def reg_exp(string: str) -> str:
    regex = r'\d+'
    try:
        return re.search(regex, string)[0]
    except:
        return None
